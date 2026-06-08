import os
import io
from datetime import datetime
from typing import List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app import models, schemas
from app.auth import get_current_user

router = APIRouter(prefix="/api/exams/{exam_id}", tags=["students"], dependencies=[Depends(get_current_user)])


def _get_exam(exam_id: int, db: Session) -> models.Exam:
    exam = db.query(models.Exam).filter(models.Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")
    return exam


# ── 列表 ─────────────────────────────────────────────────────────────────────

@router.get("/students", response_model=List[schemas.StudentOut])
def list_students(exam_id: int, db: Session = Depends(get_db)):
    return (
        db.query(models.Student)
        .filter(models.Student.exam_id == exam_id)
        .order_by(models.Student.seat_number)
        .all()
    )


# ── 生成临时学号 ──────────────────────────────────────────────────────────────

@router.post("/students/generate-temp", response_model=List[schemas.StudentOut])
def generate_temp_students(
    exam_id: int,
    body: schemas.GenerateTempRequest,
    db: Session = Depends(get_db),
):
    _get_exam(exam_id, db)
    db.query(models.Student).filter(models.Student.exam_id == exam_id).delete()
    db.flush()

    today = datetime.now().strftime("%Y%m%d")
    pad = 3 if body.count > 99 else 2
    students = []
    for seat in range(1, body.count + 1):
        sn = f"{today}{str(seat).zfill(pad)}"
        s = models.Student(
            exam_id=exam_id,
            student_number=sn,
            student_name=None,
            seat_number=seat,
            is_temp=True,
        )
        db.add(s)
        students.append(s)
    db.commit()
    for s in students:
        db.refresh(s)
    return students


# ── 导入 Excel/CSV ────────────────────────────────────────────────────────────

@router.post("/students/import", response_model=schemas.StudentImportResult)
async def import_students(
    exam_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    _get_exam(exam_id, db)
    content = await file.read()
    fname = file.filename or ""
    errors: List[str] = []
    rows = []

    if fname.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):
            rows.append((i, row))
    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                rows.append((i, dict(zip(headers, [str(v or "").strip() for v in row]))))
        except ImportError:
            raise HTTPException(status_code=400, detail="需要安装 openpyxl：pip install openpyxl")
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .csv 文件")

    db.query(models.Student).filter(models.Student.exam_id == exam_id).delete()
    db.flush()

    imported = 0
    for line_no, row in rows:
        sn = row.get("学号") or row.get("student_number", "").strip()
        if not sn:
            errors.append(f"第{line_no}行：学号为空，已跳过")
            continue
        s = models.Student(
            exam_id=exam_id,
            student_number=sn,
            student_name=row.get("姓名") or row.get("name") or None,
            class_name=row.get("班级") or row.get("class_name") or None,
            seat_number=imported + 1,
            is_temp=False,
        )
        db.add(s)
        imported += 1

    db.commit()
    return schemas.StudentImportResult(imported=imported, failed=len(errors), errors=errors)


# ── 清空 ──────────────────────────────────────────────────────────────────────

@router.delete("/students", response_model=dict)
def delete_students(exam_id: int, db: Session = Depends(get_db)):
    _get_exam(exam_id, db)
    deleted = db.query(models.Student).filter(models.Student.exam_id == exam_id).delete()
    db.commit()
    return {"deleted": deleted}


# ── 生成 PDF（后台任务） ────────────────────────────────────────────────────────

def _do_generate_pdf(exam_id: int, layout: str):
    """后台任务：生成 PDF 并保存路径到 exam.answer_sheet_path。"""
    from app.database import SessionLocal
    from app.services.sheet_generator import generate_answer_sheets
    db = SessionLocal()
    try:
        exam = db.query(models.Exam).filter(models.Exam.id == exam_id).first()
        if not exam:
            return
        students = (
            db.query(models.Student)
            .filter(models.Student.exam_id == exam_id)
            .order_by(models.Student.seat_number)
            .all()
        )
        if not students:
            return
        path = generate_answer_sheets(exam, students, layout=layout)
        exam.answer_sheet_path = path
        db.commit()
        print(f"[sheet_generator] PDF生成完成: {path}  共{len(students)}人")
    except Exception as e:
        print(f"[sheet_generator] PDF生成失败: {e}")
    finally:
        db.close()


@router.post("/students/generate-pdf", response_model=dict)
def generate_pdf(
    exam_id: int,
    background_tasks: BackgroundTasks,
    layout: str = "by_student",
    db: Session = Depends(get_db),
):
    _get_exam(exam_id, db)
    count = db.query(models.Student).filter(models.Student.exam_id == exam_id).count()
    if count == 0:
        raise HTTPException(status_code=400, detail="该考试暂无考生名单")
    background_tasks.add_task(_do_generate_pdf, exam_id, layout)
    return {"status": "generating", "student_count": count}


# ── 下载 PDF ─────────────────────────────────────────────────────────────────

@router.get("/answer-sheet/download")
def download_answer_sheet(exam_id: int, db: Session = Depends(get_db)):
    exam = _get_exam(exam_id, db)
    path = exam.answer_sheet_path
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="答题卡PDF尚未生成，请先点击生成")
    return FileResponse(
        path,
        media_type="application/pdf",
        filename=os.path.basename(path),
    )


# ── 答题卡基础 HTML ─────────────────────────────────────────────────────────────

@router.get("/answer-sheet/html")
def answer_sheet_html(exam_id: int, db: Session = Depends(get_db)):
    """返回该考试答题卡的基础 HTML（不含考生信息值和 QR 码）。"""
    from fastapi.responses import HTMLResponse
    from app.services.sheet_generator import generate_exam_html
    exam = _get_exam(exam_id, db)
    html = generate_exam_html(exam)
    return HTMLResponse(content=html)


# ── PDF 状态查询 ───────────────────────────────────────────────────────────────

@router.get("/answer-sheet/status", response_model=dict)
def answer_sheet_status(exam_id: int, db: Session = Depends(get_db)):
    exam = _get_exam(exam_id, db)
    count = db.query(models.Student).filter(models.Student.exam_id == exam_id).count()
    has_pdf = bool(exam.answer_sheet_path and os.path.exists(exam.answer_sheet_path))
    return {
        "student_count": count,
        "has_pdf": has_pdf,
        "pdf_path": os.path.basename(exam.answer_sheet_path) if has_pdf else None,
    }
