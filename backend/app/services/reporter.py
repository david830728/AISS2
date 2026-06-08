import io
from typing import Dict, Any, List
from sqlalchemy.orm import Session
from app import models


class ReportGenerator:
    def __init__(self, db: Session):
        self.db = db

    def generate_summary(self, exam: models.Exam) -> Dict[str, Any]:
        student_exams = self.db.query(models.StudentExam).filter(
            models.StudentExam.exam_id == exam.id
        ).all()

        scores = [se.total_score for se in student_exams if se.total_score is not None]
        if not scores:
            return {"exam_id": exam.id, "exam_name": exam.name, "message": "暂无成绩数据"}

        pass_line = exam.total_score * 0.6
        excellent_line = exam.total_score * 0.9

        sorted_scores = sorted(scores, reverse=True)
        top10_count = max(1, len(sorted_scores) // 10)

        return {
            "exam_id": exam.id,
            "exam_name": exam.name,
            "subject": exam.subject,
            "total_students": len(student_exams),
            "graded_count": len(scores),
            "average_score": round(sum(scores) / len(scores), 2),
            "highest_score": max(scores),
            "lowest_score": min(scores),
            "median_score": sorted_scores[len(sorted_scores) // 2],
            "pass_count": sum(1 for s in scores if s >= pass_line),
            "pass_rate": round(sum(1 for s in scores if s >= pass_line) / len(scores) * 100, 1),
            "excellent_count": sum(1 for s in scores if s >= excellent_line),
            "excellent_rate": round(sum(1 for s in scores if s >= excellent_line) / len(scores) * 100, 1),
            "top10_threshold": sorted_scores[top10_count - 1] if sorted_scores else 0,
            "score_distribution": self._score_distribution(scores, exam.total_score),
        }

    def _score_distribution(self, scores: List[float], total: float) -> List[Dict]:
        ranges = [(0, 60), (60, 70), (70, 80), (80, 90), (90, 101)]
        labels = ["不及格(<60)", "及格(60-70)", "中等(70-80)", "良好(80-90)", "优秀(90+)"]
        result = []
        for (low, high), label in zip(ranges, labels):
            count = sum(1 for s in scores if low <= (s / total * 100) < high)
            result.append({
                "label": label,
                "count": count,
                "percentage": round(count / len(scores) * 100, 1) if scores else 0,
            })
        return result

    def generate_class_analysis(self, exam: models.Exam) -> Dict[str, Any]:
        student_exams = self.db.query(models.StudentExam).filter(
            models.StudentExam.exam_id == exam.id
        ).all()

        all_scores = [se.total_score for se in student_exams if se.total_score is not None]
        pass_line = exam.total_score * 0.6

        classes: Dict[str, List[float]] = {}
        for se in student_exams:
            cn = se.class_name or "未知班级"
            if se.total_score is not None:
                classes.setdefault(cn, []).append(se.total_score)

        class_stats = []
        for cn, scores in classes.items():
            if not scores:
                continue
            class_stats.append({
                "class_name": cn,
                "student_count": len(scores),
                "average_score": round(sum(scores) / len(scores), 2),
                "max_score": max(scores),
                "min_score": min(scores),
                "pass_rate": round(sum(1 for s in scores if s >= pass_line) / len(scores) * 100, 1),
            })

        class_stats.sort(key=lambda x: x["average_score"], reverse=True)

        question_analysis = []
        for q in sorted(exam.questions, key=lambda x: x.order_index):
            answers = self.db.query(models.StudentAnswer).filter(
                models.StudentAnswer.question_id == q.id,
                models.StudentAnswer.score.isnot(None)
            ).all()
            if not answers:
                continue
            avg = sum(a.score for a in answers) / len(answers)
            question_analysis.append({
                "question_number": q.question_number,
                "question_type": q.question_type,
                "max_score": q.max_score,
                "avg_score": round(avg, 2),
                "score_rate": round(avg / q.max_score * 100, 1),
                "answer_count": len(answers),
                "zero_score_count": sum(1 for a in answers if a.score == 0),
                "full_score_count": sum(1 for a in answers if a.score == q.max_score),
            })

        dist_raw = self._score_distribution(all_scores, exam.total_score) if all_scores else []
        score_distribution = {d["label"]: d["count"] for d in dist_raw}

        return {
            "exam_id": exam.id,
            "exam_name": exam.name,
            "subject": exam.subject,
            "total_students": len(student_exams),
            "graded_students": len(all_scores),
            "avg_score": round(sum(all_scores) / len(all_scores), 2) if all_scores else 0,
            "max_score": max(all_scores) if all_scores else 0,
            "min_score": min(all_scores) if all_scores else 0,
            "pass_count": sum(1 for s in all_scores if s >= pass_line),
            "pass_rate": round(sum(1 for s in all_scores if s >= pass_line) / len(all_scores), 3) if all_scores else 0,
            "score_distribution": score_distribution,
            "class_comparison": class_stats,
            "question_analysis": question_analysis,
            "teaching_suggestions": self._generate_suggestions(question_analysis),
        }

    def _generate_suggestions(self, question_analysis: List[Dict]) -> List[str]:
        suggestions = []
        weak_questions = [q for q in question_analysis if q["score_rate"] < 50]
        if weak_questions:
            nums = ", ".join(str(q["question_number"]) for q in weak_questions[:5])
            suggestions.append(f"以下题目得分率低于50%，建议重点讲解：第{nums}题")

        zero_heavy = [q for q in question_analysis if q["answer_count"] > 0 and
                      q["zero_score_count"] / q["answer_count"] > 0.3]
        if zero_heavy:
            nums = ", ".join(str(q["question_number"]) for q in zero_heavy[:3])
            suggestions.append(f"以下题目有超过30%的学生得0分，建议加强训练：第{nums}题")

        if not suggestions:
            suggestions.append("整体表现良好，继续保持。")
        return suggestions

    def generate_excel(self, exam: models.Exam) -> io.BytesIO:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl未安装，无法生成Excel报告")

        wb = openpyxl.Workbook()

        # ── 成绩汇总 sheet ──
        ws1 = wb.active
        ws1.title = "成绩汇总"
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font_white = Font(bold=True, color="FFFFFF")
        center = Alignment(horizontal="center", vertical="center")

        questions = sorted(exam.questions, key=lambda q: q.order_index)
        headers = ["姓名", "学号", "班级"] + [f"第{q.question_number}题({q.max_score}分)" for q in questions] + ["总分", "评分状态"]

        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center

        student_exams = self.db.query(models.StudentExam).filter(
            models.StudentExam.exam_id == exam.id
        ).order_by(models.StudentExam.total_score.desc()).all()

        status_map = {
            "pending": "待评分",
            "auto_graded": "自动评分",
            "ai_graded": "AI评分",
            "manual_graded": "人工评分",
            "completed": "已完成",
        }

        for row_idx, se in enumerate(student_exams, 2):
            ws1.cell(row=row_idx, column=1, value=se.student_name or "")
            ws1.cell(row=row_idx, column=2, value=se.student_number or "")
            ws1.cell(row=row_idx, column=3, value=se.class_name or "")

            answers_map = {a.question_id: a for a in se.answers}
            for col_offset, q in enumerate(questions):
                ans = answers_map.get(q.id)
                score_val = ans.score if ans else None
                ws1.cell(row=row_idx, column=4 + col_offset, value=score_val)

            ws1.cell(row=row_idx, column=4 + len(questions), value=se.total_score)
            ws1.cell(row=row_idx, column=5 + len(questions), value=status_map.get(se.grading_status, se.grading_status))

        for col in range(1, len(headers) + 1):
            ws1.column_dimensions[get_column_letter(col)].width = 15

        # ── 成绩统计 sheet ──
        ws2 = wb.create_sheet("成绩统计")
        summary = self.generate_summary(exam)
        stat_rows = [
            ("考试名称", summary.get("exam_name", "")),
            ("科目", summary.get("subject", "")),
            ("参与人数", summary.get("total_students", 0)),
            ("已评分人数", summary.get("graded_count", 0)),
            ("平均分", summary.get("average_score", "")),
            ("最高分", summary.get("highest_score", "")),
            ("最低分", summary.get("lowest_score", "")),
            ("及格率", f"{summary.get('pass_rate', 0)}%"),
            ("优秀率", f"{summary.get('excellent_rate', 0)}%"),
        ]
        for r, (k, v) in enumerate(stat_rows, 1):
            ws2.cell(row=r, column=1, value=k).font = Font(bold=True)
            ws2.cell(row=r, column=2, value=v)

        # ── 题目分析 sheet ──
        ws3 = wb.create_sheet("题目分析")
        analysis = self.generate_class_analysis(exam)
        q_headers = ["题号", "题型", "满分", "平均分", "得分率", "作答人数", "零分人数", "满分人数"]
        for col, h in enumerate(q_headers, 1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="70AD47")
            cell.alignment = center

        type_label = {"choice": "选择题", "fill": "填空题", "subjective": "主观题"}
        for r, qa in enumerate(analysis.get("question_analysis", []), 2):
            ws3.cell(row=r, column=1, value=qa["question_number"])
            ws3.cell(row=r, column=2, value=type_label.get(qa["question_type"], qa["question_type"]))
            ws3.cell(row=r, column=3, value=qa["max_score"])
            ws3.cell(row=r, column=4, value=qa["avg_score"])
            ws3.cell(row=r, column=5, value=f"{qa['score_rate']}%")
            ws3.cell(row=r, column=6, value=qa["answer_count"])
            ws3.cell(row=r, column=7, value=qa["zero_score_count"])
            ws3.cell(row=r, column=8, value=qa["full_score_count"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
